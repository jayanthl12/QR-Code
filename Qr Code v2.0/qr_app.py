import os
import pandas as pd
import qrcode
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog, messagebox


def col_letter(n):
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def generate_qr():

    file_path = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )

    if not file_path:
        return

    try:
        excel_data = pd.ExcelFile(file_path)

        output_file = os.path.join(
            os.path.dirname(file_path),
            "QR_Output.xlsx"
        )

        wb = Workbook()
        wb.remove(wb.active)

        os.makedirs("qr_temp", exist_ok=True)

        for sheet_name in excel_data.sheet_names:

            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=None
            )

            ws = wb.create_sheet(title=sheet_name)

            for row in range(df.shape[0]):
                for col in range(df.shape[1]):

                    data = str(df.iat[row, col])

                    qr = qrcode.make(data)

                    img_path = os.path.join(
                        "qr_temp",
                        f"{sheet_name}_{row}_{col}.png"
                    )

                    qr.save(img_path)

                    img = Image(img_path)
                    img.width = 80
                    img.height = 80

                    cell = ws.cell(
                        row=row + 1,
                        column=col + 1
                    ).coordinate

                    ws.add_image(img, cell)

                    ws.row_dimensions[row + 1].height = 65
                    ws.column_dimensions[
                        col_letter(col + 1)
                    ].width = 15

        wb.save(output_file)

        messagebox.showinfo(
            "Success",
            f"Output Created:\n{output_file}"
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))


root = tk.Tk()
root.title("Production QR Excel Generator")
root.geometry("450x250")

label = tk.Label(
    root,
    text="Production QR Generator",
    font=("Arial", 18)
)
label.pack(pady=30)

btn = tk.Button(
    root,
    text="Select Excel File",
    command=generate_qr,
    width=25,
    height=3
)
btn.pack()

root.mainloop()