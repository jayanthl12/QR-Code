import os
import re
import qrcode
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import messagebox
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Image as PDFImage


def split_serial(serial):
    match = re.match(r"([A-Za-z]+)(\d+)", serial)
    if not match:
        raise ValueError(f"Invalid serial format: {serial}")
    return match.group(1), match.group(2)


def generate():
    try:
        mmyy = mmyy_entry.get()
        #version = version_entry.get()
        vendor = vendor_entry.get()
        spec = spec_entry.get()

        subsystem_lines = subsystem_text.get("1.0", tk.END).strip().split("\n")

        wb = Workbook()
        wb.remove(wb.active)

        for line in subsystem_lines:

            name, start, end,ver = [x.strip() for x in line.split(",")]

            prefix, start_num = split_serial(start)
            _, end_num = split_serial(end)

            start_num = int(start_num)
            end_num = int(end_num)

            pad = len(start.replace(prefix, ""))

            ws = wb.create_sheet(title=name)

            row = 1

            for i in range(start_num, end_num + 1):

                base_serial = prefix + str(i).zfill(pad)

                if prefix == "PEP":
                    suffixes = ["A", "B", "C", "D"]

                elif prefix in ["PID", "PDD"]:
                    suffixes = ["A", "B"]

                elif prefix == "PSP":
                    suffixes = ["A", "B", "C", "D", "E", "F"]

                else:
                    suffixes = [""]

                for suffix in suffixes:

                    serial = "Sl.No: "+base_serial + suffix

                    data = (
                        f"{vendor}\n"
                        f"{name}\n"
                        f"{mmyy} "
                        f"{serial}\n"
                        f"{spec}\n"
                        f"{ver}"
                        
                    )

                    col = ((row - 1) % 8) + 1
                    excel_row = ((row - 1) // 8) + 1

                    ws.cell(row=excel_row, column=col).value = data

                    ws.row_dimensions[excel_row].height = 90
                    ws.column_dimensions[chr(64 + col)].width = 25

                    row += 1

        #wb.save("Generated_Data.xlsx")

        #if messagebox.askyesno("QR", "Generate QR codes?"):

        qr_wb = Workbook()
        qr_wb.remove(qr_wb.active)

        os.makedirs("qr_temp", exist_ok=True)

        for sheet in wb.sheetnames:

            src = wb[sheet]
            dst = qr_wb.create_sheet(sheet)

            dst.page_setup.orientation = "landscape"
            dst.page_setup.paperSize = 9
            dst.sheet_view.zoomScale = 80

            for r in range(1, src.max_row + 1):
                for c in range(1, src.max_column + 1):

                    data = src.cell(r, c).value

                    if data:

                        img_path = f"qr_temp/{sheet}_{r}_{c}.png"

                        qrcode.make(data).save(img_path)

                        img = Image(img_path)
                        img.width = 55
                        img.height = 55

                        cell = chr(64 + c) + str(r)

                        dst.add_image(img, cell)

                        # Put serial number below QR
                        dst.row_dimensions[r].height = 45

                        dst.column_dimensions[chr(64 + c)].width = 10
                        

            qr_wb.save("QR_Output.xlsx")
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import (
                SimpleDocTemplate,
                Image as PDFImage,
                Table,
                PageBreak
            )

            pdf = SimpleDocTemplate(
                "QR_Output.pdf",
                pagesize=landscape(A4)
            )

            elements = []

            for sheet in wb.sheetnames:

                src = wb[sheet
            ]
                table_data = []

                for r in range(1, src.max_row + 1):

                    row_imgs = []

                    for c in range(1, src.max_column + 1):

                        data = src.cell(r, c).value

                        if data:
                            img_path = f"qr_temp/{sheet}_{r}_{c}.png"

                            row_imgs.append(
                                PDFImage(
                                    img_path,
                                    width=55,
                                    height=55
                                )
                            )
                        else:
                            row_imgs.append("")

                    table_data.append(row_imgs)

                table = Table(table_data)

                elements.append(table)
                elements.append(PageBreak())

            pdf.build(elements)

        messagebox.showinfo("Done", "Files generated successfully")

    except Exception as e:
        messagebox.showerror("Error", str(e))


root = tk.Tk()
root.title("Multi-Subsystem Production QR Generator")
root.geometry("700x700")


labels = [
    ("MMYY", "mmyy_entry"),
    #("Version", "version_entry"),
    ("Manufacture", "vendor_entry"),
    ("Specification", "spec_entry")
]

entries = {}

for label, var in labels:
    tk.Label(root, text=label).pack()
    e = tk.Entry(root, width=60)
    e.pack()
    entries[var] = e

globals().update(entries)

tk.Label(
    root,
    text="Subsystem,StartSerial,EndSerial,Version (one per line)"
).pack()

subsystem_text = tk.Text(root, height=20, width=80)
subsystem_text.pack()

tk.Button(
    root,
    text="Generate",
    command=generate,
    width=20,
    height=2
).pack(pady=20)

root.mainloop()